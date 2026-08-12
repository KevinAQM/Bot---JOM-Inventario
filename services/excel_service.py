import io
from datetime import datetime, date
from typing import Dict, Any, List
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from database.models import PRODUCT_CATALOG

# Orden específico de los productos para la tabla
PRODUCT_ORDER = ["V", "A", "R", "NC", "N"]

def generate_excel_report(data: Dict[str, Any]) -> io.BytesIO:
    """
    Genera un archivo Excel (.xlsx) estructurado en memoria con 3 hojas:
    1. Histórico Diario: Filas por fecha, columnas por producto (Producción y Retiro).
    2. Resumen de Inventario: Stock Inicial, Producido, Retirado y Neto Disponible.
    3. Detalle de Retiros: Lista cronológica de todos los descuentos registrados.
    """
    wb = openpyxl.Workbook()

    # Estilos comunes
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Calibri", size=11, bold=True)
    regular_font = Font(name="Calibri", size=11)
    
    fill_blue_header = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    fill_green_header = PatternFill(start_color="276A3C", end_color="276A3C", fill_type="solid")
    fill_gray_header = PatternFill(start_color="595959", end_color="595959", fill_type="solid")
    fill_total_row = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    fill_zebra = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    thin_border_side = Side(border_style="thin", color="D9D9D9")
    thick_bottom_side = Side(border_style="medium", color="000000")
    double_bottom_side = Side(border_style="double", color="000000")

    cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    header_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thick_bottom_side)
    total_border = Border(top=thin_border_side, bottom=double_bottom_side)

    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    # Extraer estructuras de datos
    productions: List = data.get("productions", [])
    withdrawals: List = data.get("withdrawals", [])
    initial_stock: Dict[str, int] = data.get("initial_stock", {})

    # Mapear producciones y retiros por fecha
    prod_map: Dict[date, Dict[str, int]] = {}
    worked_days_map: Dict[date, bool] = {}
    for p in productions:
        if p.date not in prod_map:
            prod_map[p.date] = {}
            worked_days_map[p.date] = p.is_worked_day
        prod_map[p.date][p.product_code] = p.quantity

    withd_map: Dict[date, Dict[str, int]] = {}
    for w in withdrawals:
        if w.date not in withd_map:
            withd_map[w.date] = {}
        withd_map[w.date][w.product_code] = withd_map[w.date].get(w.product_code, 0) + w.quantity

    # Obtener todas las fechas únicas ordenadas cronológicamente
    all_dates = sorted(list(set(prod_map.keys()).union(set(withd_map.keys()))))

    # ==========================================
    # HOJA 1: HISTÓRICO DIARIO
    # ==========================================
    ws1 = wb.active
    ws1.title = "Histórico Diario"
    ws1.views.sheetView[0].showGridLines = True

    # Encabezado principal
    ws1.merge_cells("A1:N1")
    title_cell = ws1["A1"]
    title_cell.value = f"REPORTE HISTÓRICO DIARIO DE PRODUCCIÓN Y RETIROS (Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')})"
    title_cell.font = Font(name="Calibri", size=14, bold=True, color="1F4E79")
    title_cell.alignment = align_left

    headers_ws1 = ["Fecha", "Día Laborado"]
    for code in PRODUCT_ORDER:
        name = PRODUCT_CATALOG.get(code, {}).get("name", code)
        headers_ws1.append(f"{name} ({code}) Prod.")
        headers_ws1.append(f"{name} ({code}) Retiro")
    headers_ws1.extend(["Total Producción", "Total Retiros"])

    ws1.append([])  # Fila 2 vacía

    # Escribir fila 3 (Encabezados de columna)
    for col_idx, text in enumerate(headers_ws1, 1):
        cell = ws1.cell(row=3, column=col_idx, value=text)
        cell.font = header_font
        cell.fill = fill_blue_header
        cell.alignment = align_center
        cell.border = header_border

    # Llenar datos de producción y retiros por fecha
    current_row = 4
    for d in all_dates:
        d_str = d.strftime("%d/%m/%Y")
        is_worked = worked_days_map.get(d, True)
        worked_text = "Sí" if is_worked else "No"

        row_vals = [d_str, worked_text]
        daily_prod_total = 0
        daily_withd_total = 0

        for code in PRODUCT_ORDER:
            p_qty = prod_map.get(d, {}).get(code, 0)
            w_qty = withd_map.get(d, {}).get(code, 0)
            row_vals.append(p_qty)
            row_vals.append(w_qty)
            daily_prod_total += p_qty
            daily_withd_total += w_qty

        row_vals.extend([daily_prod_total, daily_withd_total])

        ws1.append(row_vals)
        
        # Formato de celdas
        for c_idx in range(1, len(row_vals) + 1):
            cell = ws1.cell(row=current_row, column=c_idx)
            cell.font = regular_font
            cell.border = cell_border
            if c_idx in (1, 2):
                cell.alignment = align_center
            else:
                cell.alignment = align_right
                cell.number_format = "#,##0"

        if current_row % 2 == 1:
            for c_idx in range(1, len(row_vals) + 1):
                ws1.cell(row=current_row, column=c_idx).fill = fill_zebra

        current_row += 1

    # Fila de Totales Generales al final de la Hoja 1
    if all_dates:
        totals_row_vals = ["TOTALES", "-"]
        for code in PRODUCT_ORDER:
            tot_p = sum(prod_map.get(d, {}).get(code, 0) for d in all_dates)
            tot_w = sum(withd_map.get(d, {}).get(code, 0) for d in all_dates)
            totals_row_vals.append(tot_p)
            totals_row_vals.append(tot_w)

        grand_p = sum(sum(prod_map.get(d, {}).values()) for d in all_dates)
        grand_w = sum(sum(withd_map.get(d, {}).values()) for d in all_dates)
        totals_row_vals.extend([grand_p, grand_w])

        ws1.append(totals_row_vals)
        for c_idx in range(1, len(totals_row_vals) + 1):
            cell = ws1.cell(row=current_row, column=c_idx)
            cell.font = bold_font
            cell.fill = fill_total_row
            cell.border = total_border
            if c_idx <= 2:
                cell.alignment = align_center
            else:
                cell.alignment = align_right
                cell.number_format = "#,##0"

    # Auto-ajustar ancho de columnas
    for col in ws1.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws1.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # ==========================================
    # HOJA 2: RESUMEN DE INVENTARIO
    # ==========================================
    ws2 = wb.create_sheet(title="Resumen de Inventario")
    ws2.views.sheetView[0].showGridLines = True

    ws2.merge_cells("A1:F1")
    t2 = ws2["A1"]
    t2.value = "CONSOLIDADO DE INVENTARIO NETO ACTUAL"
    t2.font = Font(name="Calibri", size=14, bold=True, color="276A3C")
    t2.alignment = align_left

    ws2.append([])

    headers_ws2 = ["Código", "Producto", "Stock Inicial", "Total Producido", "Total Retirado", "Stock Neto Disponible"]
    ws2.append(headers_ws2)

    for col_idx, text in enumerate(headers_ws2, 1):
        cell = ws2.cell(row=3, column=col_idx)
        cell.font = header_font
        cell.fill = fill_green_header
        cell.alignment = align_center
        cell.border = header_border

    r_idx = 4
    grand_initial = 0
    grand_prod = 0
    grand_withd = 0
    grand_net = 0

    for code in PRODUCT_ORDER:
        name = PRODUCT_CATALOG.get(code, {}).get("name", code)
        init_q = initial_stock.get(code, 0)
        tot_p = sum(prod_map.get(d, {}).get(code, 0) for d in all_dates)
        tot_w = sum(withd_map.get(d, {}).get(code, 0) for d in all_dates)
        net_stock = init_q + tot_p - tot_w

        grand_initial += init_q
        grand_prod += tot_p
        grand_withd += tot_w
        grand_net += net_stock

        row_data = [code, name, init_q, tot_p, tot_w, net_stock]
        ws2.append(row_data)

        for c in range(1, 7):
            cell = ws2.cell(row=r_idx, column=c)
            cell.font = regular_font
            cell.border = cell_border
            if c in (1, 2):
                cell.alignment = align_left
            else:
                cell.alignment = align_right
                cell.number_format = "#,##0"
        r_idx += 1

    # Fila de Totales en Hoja 2
    ws2.append(["TOTALES", "Consolidado General", grand_initial, grand_prod, grand_withd, grand_net])
    for c in range(1, 7):
        cell = ws2.cell(row=r_idx, column=c)
        cell.font = bold_font
        cell.fill = fill_total_row
        cell.border = total_border
        if c in (1, 2):
            cell.alignment = align_left
        else:
            cell.alignment = align_right
            cell.number_format = "#,##0"

    for col in ws2.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws2.column_dimensions[col_letter].width = max(max_len + 4, 15)

    # ==========================================
    # HOJA 3: DETALLE DE RETIROS
    # ==========================================
    ws3 = wb.create_sheet(title="Detalle de Retiros")
    ws3.views.sheetView[0].showGridLines = True

    ws3.merge_cells("A1:G1")
    t3 = ws3["A1"]
    t3.value = "HISTORIAL DETALLADO DE SALIDAS Y RETIROS DE MERCADERÍA"
    t3.font = Font(name="Calibri", size=14, bold=True, color="595959")
    t3.alignment = align_left

    ws3.append([])

    headers_ws3 = ["ID", "Fecha", "Código", "Producto", "Cantidad", "Tipo Retiro", "Cliente / Motivo"]
    ws3.append(headers_ws3)

    for col_idx, text in enumerate(headers_ws3, 1):
        cell = ws3.cell(row=3, column=col_idx)
        cell.font = header_font
        cell.fill = fill_gray_header
        cell.alignment = align_center
        cell.border = header_border

    r_idx = 4
    for w in withdrawals:
        p_name = PRODUCT_CATALOG.get(w.product_code, {}).get("name", w.product_code)
        d_str = w.date.strftime("%d/%m/%Y")
        row_data = [w.id, d_str, w.product_code, p_name, w.quantity, w.withdrawal_type, w.customer_or_reason or "-"]
        ws3.append(row_data)

        for c in range(1, 8):
            cell = ws3.cell(row=r_idx, column=c)
            cell.font = regular_font
            cell.border = cell_border
            if c in (1, 2, 3, 6):
                cell.alignment = align_center
            elif c == 4:
                cell.alignment = align_left
            elif c == 5:
                cell.alignment = align_right
                cell.number_format = "#,##0"
            else:
                cell.alignment = align_left

        if r_idx % 2 == 1:
            for c in range(1, 8):
                ws3.cell(row=r_idx, column=c).fill = fill_zebra

        r_idx += 1

    for col in ws3.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws3.column_dimensions[col_letter].width = max(max_len + 4, 15)

    # Guardar libro en buffer BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
